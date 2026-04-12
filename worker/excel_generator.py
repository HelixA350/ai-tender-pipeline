import logging
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from worker.schemas.excel_data import (
    ExcelData,
)

logger = logging.getLogger(__name__)

DARK_BLUE_FILL = PatternFill(
    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
)
MEDIUM_BLUE_FILL = PatternFill(
    start_color="2E75B6", end_color="2E75B6", fill_type="solid"
)
LIGHT_BLUE_FILL = PatternFill(
    start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
)
VERY_LIGHT_BLUE_FILL = PatternFill(
    start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"
)

HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SECTION_FONT = Font(bold=True, size=12)
LABEL_FONT = Font(bold=True)
VALUE_FONT = Font()

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
LEFT_ALIGN_NO_wrap = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_excel_file(excel_data: Optional[ExcelData]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Закупка"

    if not excel_data:
        return save_workbook(wb)

    row = 1

    row = _write_header(ws, row, excel_data.general)

    item_columns = _get_item_columns(excel_data.items)
    row = _write_items_table(ws, row, excel_data.items, item_columns)

    row = _write_customer_section(ws, row, excel_data)
    row = _write_requirements_section(ws, row, excel_data)
    row = _write_financials_section(ws, row, excel_data)
    row = _write_dates_section(ws, row, excel_data)

    _adjust_column_width(ws, item_columns)

    return save_workbook(wb)


def save_workbook(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_header(ws, row: int, general) -> int:
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "ЗАЯВКА В ОТДЕЛ ЗАКУПОК"
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = DARK_BLUE_FILL
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    method = general.method if general and general.method else ""
    name = general.name if general and general.name else ""
    cell.value = f"{method}: {name}" if method and name else (method or name or "")
    cell.font = Font(color="FFFFFF", size=12)
    cell.alignment = CENTER_ALIGN
    cell.fill = MEDIUM_BLUE_FILL
    row += 1

    if general and general.notes:
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws[f"A{row}"]
        cell.value = f"*{general.notes}"
        cell.font = Font(size=11, italic=True)
        cell.alignment = CENTER_ALIGN
        cell.fill = VERY_LIGHT_BLUE_FILL
        row += 1
    else:
        row += 1

    row += 1
    return row


def _get_item_columns(items) -> dict:
    if not items:
        return {}

    field_map = {
        "position": "№",
        "name": "Наименование",
        "article": "Артикул",
        "manufacturer": "Производитель",
        "qty": "Кол-во",
        "unit": "Ед.изм",
        "unit_price": "Цена",
        "currency": "Валюта",
        "delivery_date": "Срок поставки",
        "delivery_location": "Место поставки",
        "analog_allowed": "Аналог",
        "npp": "Код НПП",
        "category": "Категория",
        "original_reference": "Ссылка",
        "source": "Файл",
        "notes": "Примечания",
        "linked_service": "Услуга",
    }

    present_columns = {}
    for field, header in field_map.items():
        for item in items:
            value = getattr(item, field, None)
            if value is not None:
                present_columns[field] = header
                break

    return present_columns


def _write_items_table(ws, row: int, items, columns: dict) -> int:
    if not items or not columns:
        row += 1
        return row

    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "📋  ПЕРЕЧЕНЬ ПОЗИЦИЙ ДЛЯ ЗАКУПКИ"
    cell.font = SECTION_FONT
    cell.alignment = LEFT_ALIGN_NO_wrap
    cell.fill = LIGHT_BLUE_FILL
    row += 1

    num_cols = len(columns)

    for idx, (field, header) in enumerate(columns.items()):
        col = idx + 1
        cell = ws.cell(row, col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = MEDIUM_BLUE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    row += 1

    for idx, item in enumerate(items):
        fill = LIGHT_BLUE_FILL if idx % 2 == 0 else PatternFill()
        for jdx, (field, header) in enumerate(columns.items()):
            col = jdx + 1
            cell = ws.cell(row, col)
            value = getattr(item, field, None)
            if field == "analog_allowed":
                cell.value = "Да" if value else "Нет"
            elif value is not None:
                cell.value = value
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = CENTER_ALIGN
            if field == "name":
                cell.alignment = LEFT_ALIGN

        row += 1

    row += 1
    return row


def _write_section_header(ws, row: int, title: str) -> int:
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = SECTION_FONT
    cell.alignment = LEFT_ALIGN_NO_wrap
    cell.fill = LIGHT_BLUE_FILL
    row += 1
    return row


def _write_field(ws, row: int, label: str, value) -> int:
    if value is None or value == "":
        return row

    cell_label = ws.cell(row, 1)
    cell_label.value = label
    cell_label.font = LABEL_FONT
    cell_label.alignment = LEFT_ALIGN
    cell_label.fill = VERY_LIGHT_BLUE_FILL

    ws.merge_cells(f"B{row}:D{row}")
    cell_value = ws.cell(row, 2)
    cell_value.value = (
        str(value) if not isinstance(value, bool) else ("Да" if value else "Нет")
    )
    cell_value.font = VALUE_FONT
    cell_value.alignment = LEFT_ALIGN

    row += 1
    return row


def _write_customer_section(ws, row: int, excel_data: ExcelData) -> int:
    if not excel_data.customer:
        return row

    row = _write_section_header(ws, row, "🏭 Заказчик")
    customer = excel_data.customer

    row = _write_field(ws, row, "Наименование", customer.name)
    row = _write_field(ws, row, "Полное наименование", customer.full_name)
    row = _write_field(ws, row, "ИНН", customer.inn)
    row = _write_field(ws, row, "КПП", customer.kpp)
    row = _write_field(ws, row, "Адрес", customer.address)
    row = _write_field(ws, row, "Закупочная организация", customer.procurement_org)
    row = _write_field(ws, row, "Группа закупок", customer.procurement_group)
    row = _write_field(ws, row, "Примечания", customer.notes)

    row += 1
    return row


def _write_requirements_section(ws, row: int, excel_data: ExcelData) -> int:
    if not excel_data.requirements:
        return row

    row = _write_section_header(ws, row, "📦 Требования к товару")
    req = excel_data.requirements

    row = _write_field(ws, row, "Состояние", req.condition)
    row = _write_field(ws, row, "Гарантия (мес)", req.warranty_months)
    row = _write_field(ws, row, "Точка отсчёта гарантии", req.warranty_start)
    row = _write_field(ws, row, "Аналоги допускаются", req.analog_allowed)
    row = _write_field(ws, row, "Правила замены", req.analog_rules)
    row = _write_field(
        ws, row, "Импортозамещение требуется", req.import_substitution_required
    )
    row = _write_field(ws, row, "Ограничения по происхождению", req.origin_restrictions)
    row = _write_field(ws, row, "Примечания", req.notes)

    row += 1
    return row


def _write_financials_section(ws, row: int, excel_data: ExcelData) -> int:
    if not excel_data.financials:
        return row

    row = _write_section_header(ws, row, "💰 Финансовые условия")
    fin = excel_data.financials

    row = _write_field(ws, row, "НМЦ", fin.nmck)
    row = _write_field(ws, row, "Валюта", fin.base_currency)
    row = _write_field(ws, row, "НДС", fin.vat_rate)
    row = _write_field(ws, row, "Цены с НДС", fin.prices_include_vat)
    row = _write_field(ws, row, "Шаг аукциона", fin.auction_step)

    if fin.bid_security and fin.bid_security.amount:
        row = _write_field(
            ws,
            row,
            "Обеспечение заявки",
            f"{fin.bid_security.form}: {fin.bid_security.amount}",
        )

    if fin.contract_security and fin.contract_security.amount:
        row = _write_field(
            ws,
            row,
            "Обеспечение контракта",
            f"{fin.contract_security.form}: {fin.contract_security.amount}",
        )

    if fin.payment_terms:
        pt = fin.payment_terms
        parts = []
        if pt.type:
            parts.append(pt.type)
        if pt.advance_pct:
            parts.append(f"аванс {pt.advance_pct}%")
        if pt.days_min and pt.days_max:
            parts.append(f"срок {pt.days_min}-{pt.days_max} дн")
        if parts:
            row = _write_field(ws, row, "Условия оплаты", ", ".join(parts))

    if fin.incoterms:
        inc = fin.incoterms
        parts = []
        if inc.primary:
            parts.append(inc.primary)
        if inc.location:
            parts.append(inc.location)
        if parts:
            row = _write_field(ws, row, "Incoterms", " ".join(parts))

    if fin.penalties:
        pen = fin.penalties
        parts = []
        if pen.late_delivery_pct:
            parts.append(f"{pen.late_delivery_pct}% за день просрочки")
        if pen.max_penalty_pct:
            parts.append(f"макс {pen.max_penalty_pct}%")
        if parts:
            row = _write_field(ws, row, "Пени", ", ".join(parts))

    row = _write_field(ws, row, "Примечания", fin.notes)

    row += 1
    return row


def _write_dates_section(ws, row: int, excel_data: ExcelData) -> int:
    if not excel_data.dates:
        return row

    row = _write_section_header(ws, row, "📅 Даты и сроки")
    dates = excel_data.dates

    row = _write_field(ws, row, "Дата публикации", dates.publication_date)
    row = _write_field(ws, row, "Срок подачи заявки", dates.submission_deadline)
    row = _write_field(ws, row, "Время подачи", dates.submission_time)
    row = _write_field(ws, row, "Дата вскрытия", dates.opening_date)
    row = _write_field(ws, row, "Дата подведения итогов", dates.results_date)
    row = _write_field(ws, row, "Начало поставки", dates.delivery_start)
    row = _write_field(ws, row, "Окончание поставки", dates.delivery_end)
    row = _write_field(ws, row, "Досрочная поставка", dates.early_delivery_allowed)
    row = _write_field(ws, row, "Примечания", dates.notes)

    row += 1
    return row


def _adjust_column_width(ws, item_columns: dict):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    for col_idx in range(5, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
